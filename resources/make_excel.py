from openpyxl import Workbook, load_workbook
from contextlib import closing

def make_excel(file_name):
    with closing(Workbook()) as wb:
        wb.save(file_name)

def add_value(file_name, parameter, valuetostore):
    with closing(load_workbook(filename=file_name)) as wb:
        sheet = wb.active

        c1 = sheet.cell(row = 1, column = 1)
        c1.value = "Parameter"
        sheet.max_row +1

        c2 = sheet.cell(row= 1 , column = 2)
        c2.value = "Value"


        c1 = sheet.cell(row =  sheet.max_row +1, column = 1)
        c1.value = parameter
  
        c2 = sheet.cell(row=  sheet.max_row , column = 2)
        c2.value = valuetostore
        wb.save(file_name)